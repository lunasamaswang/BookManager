import re

from openpyxl import load_workbook

from app.services.book_service import add_book, get_books


class ExcelImportError(Exception):
    """Excel 文件无法正常导入时抛出的业务异常。"""


def normalize_title(title):
    """清理书名，并生成用于判断重复的比较值。"""
    return str(title).strip().casefold()


def get_next_book_number(books):
    """查找现有 BOOK 编号中的最大数字，并返回下一个数字。"""
    max_number = 0
    for book in books:
        nfc_id = (book.get("nfc_id") or "").strip()
        match = re.fullmatch(r"BOOK(\d+)", nfc_id, flags=re.IGNORECASE)
        if match:
            max_number = max(max_number, int(match.group(1)))
    return max_number + 1


def generate_nfc_id(next_number, used_nfc_ids):
    """生成一个未被使用的递增 NFC 编号。"""
    while True:
        nfc_id = f"BOOK{next_number:06d}"
        next_number += 1
        if nfc_id.casefold() not in used_nfc_ids:
            used_nfc_ids.add(nfc_id.casefold())
            return nfc_id, next_number


def import_books_from_excel(file_path):
    """读取 Excel 第一列并逐行导入图书，返回导入统计结果。"""
    result = {
        "imported": 0,
        "empty_rows": 0,
        "duplicate_titles": 0,
        "failed": 0,
        "errors": [],
    }

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as error:
        raise ExcelImportError(f"无法读取 Excel 文件：{error}") from error

    try:
        worksheet = workbook.worksheets[0]
        header_value = worksheet.cell(row=1, column=1).value
        header = "" if header_value is None else str(header_value).strip().casefold()
        if header not in {"书名", "title"}:
            raise ExcelImportError(
                "Excel 第一行第一列必须是“书名”或“title”。"
            )

        existing_books = get_books()
        known_titles = {
            normalize_title(book["title"])
            for book in existing_books
            if book.get("title") is not None
        }
        used_nfc_ids = {
            str(book["nfc_id"]).strip().casefold()
            for book in existing_books
            if book.get("nfc_id")
        }
        next_number = get_next_book_number(existing_books)

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                min_col=1,
                max_col=1,
                values_only=True,
            ),
            start=2,
        ):
            cell_value = row[0]
            title = "" if cell_value is None else str(cell_value).strip()

            if not title:
                result["empty_rows"] += 1
                continue

            normalized_title = normalize_title(title)
            if normalized_title in known_titles:
                result["duplicate_titles"] += 1
                continue

            nfc_id, next_number = generate_nfc_id(next_number, used_nfc_ids)
            try:
                add_book(
                    nfc_id=nfc_id,
                    title=title,
                    author="",
                    category="未分类",
                    location="未设置",
                    status="未借出",
                )
                known_titles.add(normalized_title)
                result["imported"] += 1
            except Exception as error:
                result["failed"] += 1
                result["errors"].append(
                    {
                        "row": row_number,
                        "title": title,
                        "reason": str(error),
                    }
                )

        return result
    finally:
        workbook.close()
